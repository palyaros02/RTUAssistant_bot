from bs4 import BeautifulSoup
import requests
import csv
import lxml
import openpyxl
import re
import json

# ==== Parsing hrefs ====
URL = "https://www.mirea.ru/schedule/"
HOST = "https://www.mirea.ru/"
HEADERS = {
    'accept: text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
    'user-Agent: Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:86.0) Gecko/20100101 Firefox/86.0'
}

def get_src(url):
    req = requests.get(url)
    src = req.text
    return src

filtered = []
def get_content(src):
    soup = BeautifulSoup(src, "lxml")
    items = soup.find_all(class_="uk-link-toggle")
    hrefs = []
    for item in items:
        item_href = item.get("href").replace(" ", "%20")
        hrefs.append(item_href)
    with open("./hrefs.txt", "w", newline='', encoding="utf-8") as file:
        writer = csv.writer(file, delimiter=';')
        for item in hrefs:
            if item[-4:] == "xlsx" \
                    and item[-23:-20] == "ИИТ":  # пока только для ИТ
                filtered.append(item)
                writer.writerow([item])

get_content(get_src(URL))
# ============

# ==== Downloading files ====
def download(hrefs):
    for href in hrefs:
        with open(r'./xlsx/' + href[-23:], 'wb') as f:
            data = requests.get(href)
            f.write(data.content)
download(filtered)
# ============================

res = dict()
for href in filtered:
    wb = openpyxl.load_workbook('./xlsx/'+href[-23:])
    sheet = wb[wb.sheetnames[0]]
    for i in range(1, sheet.max_column):
        group = sheet.cell(row=2, column=i).value
        if group is not None:
            if re.match(r'[А-Я]{4}-\d\d-\d\d', str(group)) is not None:
                r = 3
                days = {}
                for day in range(1, 7):
                    pairs = []
                    for shed in range(1, 13):
                        content = sheet.cell(row=12*(day-1) + r + shed, column=i).value
                        if content is not None:
                            is_even = (shed % 2 == 0)
                            pair_no = shed // 2 if is_even else shed // 2 + 1
                            type = sheet.cell(row=12*(day-1) + r + shed, column=i+1).value
                            teacher = sheet.cell(row=12*(day-1) + r + shed, column=i+2).value
                            aud = sheet.cell(row=12*(day-1) + r + shed, column=i+3).value
                            pairs.append((is_even, pair_no, content, type, teacher, aud))
                    days[day] = pairs.copy()
                res[group] = days.copy()
with open("./schedule.json", 'w', encoding="utf-8") as f:
    f.write(json.dumps(res, ensure_ascii=False, separators=(',', ':')))
